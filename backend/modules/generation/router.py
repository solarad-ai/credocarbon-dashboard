"""
Generation Data API Router
Endpoints for file upload, data processing, and credit estimation
"""
import os
import hashlib
import csv
import io
from datetime import datetime
from typing import List, Optional, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.orm import Session

from backend.core.database import get_db
from backend.modules.auth.dependencies import get_current_user
from backend.core.models import User, Project

from .models import UploadedFile, DatasetMapping, GenerationTimeseries, CreditEstimation
from .schemas import (
    FileUploadResponse,
    FilePreviewResponse,
    ColumnInfo,
    DatasetMappingCreate,
    DatasetMappingResponse,
    MappingValidationResult,
    MethodologyInfo,
    MethodologyListResponse,
    GridEFInfo,
    GridEFListResponse,
    EstimationRequest,
    EstimationResponse,
    MonthlyBreakdown,
    AnnualBreakdown,
    ProcessingStatusResponse,
)
from .methodologies.registry import MethodologyRegistry
from .grid_ef_database import get_grid_ef, get_all_grid_efs, get_countries_list
from .services.credit_calculator import CreditCalculator

router = APIRouter(prefix="/generation", tags=["Generation Data"])

# File upload directory - use /tmp for Cloud Run compatibility
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "/tmp/uploads/generation")
try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
except PermissionError:
    # Fallback for restricted environments
    UPLOAD_DIR = "/tmp/uploads/generation"
    os.makedirs(UPLOAD_DIR, exist_ok=True)


# ============ File Upload Endpoints ============

@router.post("/upload", response_model=FileUploadResponse)
async def upload_file(
    project_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload a generation data file (CSV or Excel).
    
    Accepts CSV, XLSX, and XLS files containing generation data.
    Returns file metadata and detected columns for mapping.
    """
    # Validate file type
    allowed_types = [
        "text/csv",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]
    allowed_extensions = [".csv", ".xlsx", ".xls"]
    
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
        )
    
    # Verify project exists and user has access
    project = db.query(Project).filter(
        Project.id == project_id,
        Project.developer_id == current_user.id
    ).first()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Read file content
    content = await file.read()
    file_size = len(content)
    
    # Calculate checksum
    checksum = hashlib.sha256(content).hexdigest()
    
    # Generate storage path
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    storage_filename = f"{project_id}_{timestamp}_{file.filename}"
    storage_path = os.path.join(UPLOAD_DIR, storage_filename)
    
    # Save file
    with open(storage_path, "wb") as f:
        f.write(content)
    
    # Parse file to detect columns
    detected_columns = None
    row_count = 0
    column_count = 0
    
    try:
        if file_ext == ".csv":
            detected_columns, row_count, column_count = _parse_csv_columns(content)
        elif file_ext in [".xlsx", ".xls"]:
            # For Excel, parse after saving since openpyxl needs a file path
            pass  # Will be parsed in preview endpoint or below
    except Exception as e:
        # File saved but couldn't parse - will handle in preview
        pass
    
    # Create database record
    uploaded_file = UploadedFile(
        project_id=project_id,
        original_filename=file.filename,
        mime_type=file.content_type or "application/octet-stream",
        storage_uri=storage_path,
        file_size_bytes=file_size,
        checksum=checksum,
        row_count=row_count,
        column_count=column_count,
        detected_columns=detected_columns,
        uploaded_by=current_user.id,
        status="parsed" if detected_columns else "pending"
    )
    
    db.add(uploaded_file)
    db.commit()
    db.refresh(uploaded_file)
    
    return FileUploadResponse(
        id=uploaded_file.id,
        original_filename=uploaded_file.original_filename,
        mime_type=uploaded_file.mime_type,
        file_size_bytes=uploaded_file.file_size_bytes,
        status=uploaded_file.status,
        detected_columns=uploaded_file.detected_columns,
        uploaded_at=uploaded_file.uploaded_at
    )


def _parse_csv_columns(content: bytes) -> tuple:
    """Parse CSV to detect columns and infer types."""
    try:
        # Try to decode as UTF-8
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    
    if not rows:
        return None, 0, 0
    
    headers = rows[0]
    data_rows = rows[1:100]  # Sample first 100 rows
    
    columns = []
    for i, header in enumerate(headers):
        sample_values = [row[i] if i < len(row) else None for row in data_rows[:5]]
        inferred_type = _infer_column_type(sample_values)
        null_count = sum(1 for row in data_rows if i >= len(row) or not row[i])
        
        columns.append({
            "name": header,
            "inferred_type": inferred_type,
            "sample_values": sample_values,
            "null_count": null_count
        })
    
    return columns, len(rows) - 1, len(headers)


def _infer_column_type(sample_values: List[Any]) -> str:
    """Infer column type from sample values."""
    for value in sample_values:
        if not value:
            continue
        
        # Try datetime
        try:
            from dateutil import parser
            parser.parse(str(value))
            return "datetime"
        except:
            pass
        
        # Try numeric
        try:
            float(str(value).replace(",", ""))
            return "numeric"
        except:
            pass
    
    return "string"


@router.get("/{file_id}/preview", response_model=FilePreviewResponse)
async def get_file_preview(
    file_id: int,
    rows: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get preview of uploaded file.
    
    Returns first N rows and column metadata for mapping configuration.
    """
    uploaded_file = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    
    if not uploaded_file:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Verify user has access
    project = db.query(Project).filter(
        Project.id == uploaded_file.project_id,
        Project.developer_id == current_user.id
    ).first()
    
    if not project:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Read and parse file
    with open(uploaded_file.storage_uri, "rb") as f:
        content = f.read()
    
    filename_lower = uploaded_file.original_filename.lower()
    if filename_lower.endswith(".csv"):
        columns, preview_rows, total_rows = _parse_csv_preview(content, rows)
    elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        columns, preview_rows, total_rows = _parse_excel_preview(uploaded_file.storage_uri, rows)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format")
    
    # Update file status if it was pending
    if uploaded_file.status == "pending":
        uploaded_file.status = "parsed"
        uploaded_file.detected_columns = columns
        uploaded_file.row_count = total_rows
        uploaded_file.column_count = len(columns)
        db.commit()
    
    return FilePreviewResponse(
        file_id=file_id,
        filename=uploaded_file.original_filename,
        columns=[ColumnInfo(**col) for col in columns],
        preview_rows=preview_rows,
        total_rows=total_rows,
        total_columns=len(columns)
    )


def _parse_csv_preview(content: bytes, num_rows: int) -> tuple:
    """Parse CSV for preview display."""
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    
    if not rows:
        return [], [], 0
    
    headers = rows[0]
    data_rows = rows[1:num_rows + 1]
    all_data_rows = rows[1:]
    
    columns = []
    for i, header in enumerate(headers):
        sample_values = [row[i] if i < len(row) else None for row in data_rows[:5]]
        null_count = sum(1 for row in all_data_rows if i >= len(row) or not row[i])
        
        columns.append({
            "name": header,
            "inferred_type": _infer_column_type(sample_values),
            "sample_values": sample_values,
            "null_count": null_count
        })
    
    return columns, data_rows, len(all_data_rows)


def _parse_excel_preview(file_path: str, num_rows: int) -> tuple:
    """Parse Excel file for preview display using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required for Excel parsing. Install with: pip install openpyxl"
        )
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > num_rows + 1:  # Header + preview rows
                break
            # Convert None values to empty strings for consistency
            rows.append([str(cell) if cell is not None else "" for cell in row])
        
        if not rows:
            wb.close()
            return [], [], 0
        
        headers = rows[0]
        data_rows = rows[1:num_rows + 1]
        
        # Count total rows (approximate for large files)
        total_rows = ws.max_row - 1 if ws.max_row else 0
        
        columns = []
        for i, header in enumerate(headers):
            sample_values = [row[i] if i < len(row) else "" for row in data_rows[:5]]
            null_count = sum(1 for row in data_rows if i >= len(row) or not row[i])
            
            columns.append({
                "name": header if header else f"Column_{i+1}",
                "inferred_type": _infer_column_type(sample_values),
                "sample_values": sample_values,
                "null_count": null_count
            })
        
        wb.close()
        return columns, data_rows, total_rows
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")


# ============ Column Mapping Endpoints ============

@router.post("/{file_id}/mapping", response_model=DatasetMappingResponse)
async def save_column_mapping(
    file_id: int,
    mapping: DatasetMappingCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Save column mapping configuration for a file.
    
    Specifies which columns contain timestamp and generation data,
    along with unit and frequency information.
    """
    uploaded_file = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    
    if not uploaded_file:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Verify user has access
    project = db.query(Project).filter(
        Project.id == uploaded_file.project_id,
        Project.developer_id == current_user.id
    ).first()
    
    if not project:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Validate mapping against file columns
    if uploaded_file.detected_columns:
        column_names = [col["name"] for col in uploaded_file.detected_columns]
        if mapping.timestamp_column not in column_names:
            raise HTTPException(
                status_code=400,
                detail=f"Timestamp column '{mapping.timestamp_column}' not found in file"
            )
        if mapping.value_column not in column_names:
            raise HTTPException(
                status_code=400,
                detail=f"Value column '{mapping.value_column}' not found in file"
            )
    
    # Check if mapping already exists
    existing = db.query(DatasetMapping).filter(DatasetMapping.file_id == file_id).first()
    if existing:
        # Update existing
        for key, value in mapping.dict().items():
            setattr(existing, key, value)
        db.commit()
        db.refresh(existing)
        dataset_mapping = existing
    else:
        # Create new
        dataset_mapping = DatasetMapping(
            file_id=file_id,
            **mapping.dict()
        )
        db.add(dataset_mapping)
        db.commit()
        db.refresh(dataset_mapping)
    
    # Update file status
    uploaded_file.status = "mapped"
    db.commit()
    
    return DatasetMappingResponse.from_orm(dataset_mapping)


@router.post("/{file_id}/validate-mapping", response_model=MappingValidationResult)
async def validate_mapping(
    file_id: int,
    mapping: DatasetMappingCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Validate column mapping without saving.
    
    Returns validation results including warnings and a sample conversion.
    """
    uploaded_file = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    
    if not uploaded_file:
        raise HTTPException(status_code=404, detail="File not found")
    
    errors = []
    warnings = []
    sample_conversion = None
    detected_frequency = None
    
    # Check columns exist
    if uploaded_file.detected_columns:
        column_names = [col["name"] for col in uploaded_file.detected_columns]
        
        if mapping.timestamp_column not in column_names:
            errors.append(f"Timestamp column '{mapping.timestamp_column}' not found")
        
        if mapping.value_column not in column_names:
            errors.append(f"Value column '{mapping.value_column}' not found")
        
        # Try to detect frequency from sample timestamps
        # (Would need to parse actual file data for this)
    
    # Unit validation
    if mapping.unit not in ["kW", "MW", "kWh", "MWh"]:
        errors.append(f"Invalid unit: {mapping.unit}")
    
    # Semantics validation
    if mapping.value_semantics not in ["POWER", "ENERGY_PER_INTERVAL"]:
        errors.append(f"Invalid value_semantics: {mapping.value_semantics}")
    
    # Power vs Energy warning
    if mapping.unit in ["kW", "MW"] and mapping.value_semantics == "ENERGY_PER_INTERVAL":
        warnings.append(
            f"Unit is '{mapping.unit}' (power) but semantics is 'ENERGY_PER_INTERVAL'. "
            "This seems inconsistent."
        )
    
    if mapping.unit in ["kWh", "MWh"] and mapping.value_semantics == "POWER":
        warnings.append(
            f"Unit is '{mapping.unit}' (energy) but semantics is 'POWER'. "
            "This seems inconsistent."
        )
    
    # Sample conversion
    if not errors:
        sample_conversion = {
            "original_value": 1000,
            "original_unit": mapping.unit,
            "converted_value": _convert_to_mwh(1000, mapping.unit, mapping.value_semantics, mapping.frequency_seconds),
            "converted_unit": "MWh",
            "interval_seconds": mapping.frequency_seconds
        }
    
    return MappingValidationResult(
        valid=len(errors) == 0,
        warnings=warnings,
        errors=errors,
        sample_conversion=sample_conversion,
        detected_frequency=detected_frequency
    )


def _convert_to_mwh(value: float, unit: str, semantics: str, frequency_seconds: int) -> float:
    """Convert value to MWh based on unit and semantics."""
    if semantics == "ENERGY_PER_INTERVAL":
        # Value is already energy
        if unit == "kWh":
            return value / 1000
        elif unit == "MWh":
            return value
        elif unit == "kW":
            # Treat as kWh if semantics says energy
            return value / 1000
        elif unit == "MW":
            return value
    else:
        # Value is power, need to convert to energy
        hours = frequency_seconds / 3600
        if unit == "kW":
            return (value / 1000) * hours
        elif unit == "MW":
            return value * hours
        elif unit == "kWh":
            # Already energy, just convert
            return value / 1000
        elif unit == "MWh":
            return value
    
    return value


# ============ Methodology Endpoints ============

@router.get("/methodologies", response_model=MethodologyListResponse)
async def list_methodologies(
    project_type: Optional[str] = None,
    registry: Optional[str] = None
):
    """
    List available carbon credit methodologies.
    
    Can be filtered by project type (solar, wind, hydro, biogas, etc.)
    or by registry (CDM, VERRA, GOLD_STANDARD, GCC).
    """
    if project_type:
        methodologies = MethodologyRegistry.list_for_project_type(project_type)
    elif registry:
        methodologies = MethodologyRegistry.list_for_registry(registry)
    else:
        methodologies = MethodologyRegistry.list_all()
    
    return MethodologyListResponse(
        methodologies=[MethodologyInfo(**m) for m in methodologies]
    )


@router.get("/methodologies/{methodology_id}")
async def get_methodology(methodology_id: str):
    """Get details of a specific methodology."""
    try:
        methodology = MethodologyRegistry.get(methodology_id)
        return methodology.get_info()
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# ============ Grid Emission Factor Endpoints ============

@router.get("/grid-ef", response_model=GridEFListResponse)
async def list_grid_emission_factors(
    country_code: Optional[str] = None
):
    """
    List grid emission factors.
    
    Returns emission factors for all countries, or a specific country if provided.
    """
    if country_code:
        ef = get_grid_ef(country_code)
        if not ef:
            raise HTTPException(
                status_code=404,
                detail=f"No emission factor data for country: {country_code}"
            )
        return GridEFListResponse(emission_factors=[GridEFInfo(
            country_code=ef.country_code,
            country_name=ef.country_name,
            region_code=ef.region_code,
            region_name=ef.region_name,
            combined_margin=ef.combined_margin,
            operating_margin=ef.operating_margin,
            build_margin=ef.build_margin,
            source_name=ef.source_name,
            data_year=ef.data_year,
            source_url=ef.source_url
        )])
    
    all_efs = get_all_grid_efs()
    return GridEFListResponse(emission_factors=[
        GridEFInfo(
            country_code=ef.country_code,
            country_name=ef.country_name,
            region_code=ef.region_code,
            region_name=ef.region_name,
            combined_margin=ef.combined_margin,
            operating_margin=ef.operating_margin,
            build_margin=ef.build_margin,
            source_name=ef.source_name,
            data_year=ef.data_year,
            source_url=ef.source_url
        )
        for ef in all_efs
    ])


@router.get("/grid-ef/countries")
async def list_countries():
    """List all countries with available emission factor data."""
    return get_countries_list()


# ============ Credit Estimation Endpoints ============

@router.post("/estimate", response_model=EstimationResponse)
async def estimate_credits(
    request: EstimationRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Calculate carbon credit estimation for a project.
    
    Uses uploaded generation data and selected methodology to calculate
    emission reductions. Results are saved to the database.
    """
    # Verify project exists and user has access
    project = db.query(Project).filter(
        Project.id == request.project_id,
        Project.developer_id == current_user.id
    ).first()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get generation data
    timeseries = db.query(GenerationTimeseries).filter(
        GenerationTimeseries.project_id == request.project_id
    )
    
    if request.period_start:
        timeseries = timeseries.filter(GenerationTimeseries.ts_utc >= request.period_start)
    if request.period_end:
        timeseries = timeseries.filter(GenerationTimeseries.ts_utc <= request.period_end)
    
    timeseries = timeseries.all()
    
    # If no timeseries data, check for uploaded files with mappings
    if not timeseries:
        # Try to get data from wizard_data or use sample calculation
        raise HTTPException(
            status_code=400,
            detail="No generation data found. Please upload and process data first."
        )
    
    # Prepare generation data for calculator
    generation_data = [
        {"timestamp": ts.ts_utc, "energy_mwh": float(ts.energy_mwh)}
        for ts in timeseries
    ]
    
    # Run calculation
    try:
        calculator = CreditCalculator(request.methodology_id)
        result = calculator.calculate(
            generation_data=generation_data,
            country_code=request.country_code,
            project_type=project.project_type,
            ef_override=request.ef_value,
            additional_inputs=request.additional_inputs
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # Save estimation to database
    estimation = CreditEstimation(
        project_id=request.project_id,
        methodology_id=request.methodology_id,
        registry=result["registry"],
        country_code=request.country_code,
        grid_ef_value=result["ef_value"],
        grid_ef_source=result["ef_source"],
        grid_ef_year=result["ef_year"],
        total_generation_mwh=result["total_generation_mwh"],
        total_er_tco2e=result["total_er_tco2e"],
        baseline_emissions_tco2e=result["baseline_emissions_tco2e"],
        project_emissions_tco2e=result["project_emissions_tco2e"],
        leakage_tco2e=result["leakage_tco2e"],
        monthly_breakdown=result["monthly_breakdown"],
        annual_breakdown=result["annual_breakdown"],
        calculation_inputs=request.additional_inputs,
        assumptions=result["assumptions"],
        period_start=request.period_start,
        period_end=request.period_end,
        created_by=current_user.id
    )
    
    db.add(estimation)
    db.commit()
    db.refresh(estimation)
    
    return EstimationResponse(
        id=estimation.id,
        project_id=estimation.project_id,
        methodology_id=estimation.methodology_id,
        registry=estimation.registry,
        total_generation_mwh=float(estimation.total_generation_mwh),
        total_er_tco2e=float(estimation.total_er_tco2e),
        baseline_emissions_tco2e=float(estimation.baseline_emissions_tco2e),
        project_emissions_tco2e=float(estimation.project_emissions_tco2e),
        leakage_tco2e=float(estimation.leakage_tco2e),
        country_code=estimation.country_code,
        ef_value=float(estimation.grid_ef_value),
        ef_source=estimation.grid_ef_source,
        ef_year=estimation.grid_ef_year,
        monthly_breakdown=[MonthlyBreakdown(**m) for m in estimation.monthly_breakdown or []],
        annual_breakdown=[AnnualBreakdown(**a) for a in estimation.annual_breakdown or []],
        calculation_date=estimation.calculation_date,
        assumptions=estimation.assumptions
    )


@router.post("/quick-estimate")
async def quick_estimate(
    generation_mwh: float = Form(...),
    country_code: str = Form(...),
    project_type: str = Form("solar"),
    methodology_id: str = Form("CDM_AMS_ID"),
):
    """
    Quick estimation without saving to database.
    
    Useful for preview calculations before formal estimation.
    """
    try:
        calculator = CreditCalculator(methodology_id)
        result = calculator.calculate_simple(
            total_generation_mwh=generation_mwh,
            country_code=country_code,
            project_type=project_type
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/estimations/{project_id}")
async def list_estimations(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all credit estimations for a project."""
    # Verify access
    project = db.query(Project).filter(
        Project.id == project_id,
        Project.developer_id == current_user.id
    ).first()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    estimations = db.query(CreditEstimation).filter(
        CreditEstimation.project_id == project_id
    ).order_by(CreditEstimation.calculation_date.desc()).all()
    
    return [
        {
            "id": e.id,
            "methodology_id": e.methodology_id,
            "registry": e.registry,
            "total_er_tco2e": float(e.total_er_tco2e),
            "calculation_date": e.calculation_date,
            "ef_value": float(e.grid_ef_value),
            "country_code": e.country_code
        }
        for e in estimations
    ]
